# utils/export.py
# Builds a styled Excel report for the address comparison result.

import io
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


def build_excel(extracted_full: str, db_addr: str, row_status: str) -> bytes:
    """Create a colour-coded Excel workbook comparing extracted vs DB address.

    Args:
        extracted_full: Address string extracted from the QR code.
        db_addr:        Matching address found in the database.
        row_status:     "Match" or "Mismatch".

    Returns:
        Raw bytes of the .xlsx file.
    """
    is_match   = row_status == "Match"
    data_fill  = PatternFill("solid", fgColor="C6EFCE" if is_match else "FFC7CE")
    data_font  = Font(color="276221" if is_match else "9C0006", bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    hdr_font   = Font(bold=True, size=12)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df = pd.DataFrame([{
            "Extracted Address": extracted_full,
            "DB Address":        db_addr,
            "Status":            row_status,
        }])
        df.to_excel(writer, index=False, sheet_name="Address Comparison")
        ws = writer.sheets["Address Comparison"]

        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 16

        for cell in ws[1]:
            cell.font      = hdr_font
            cell.alignment = hdr_align

        ws.row_dimensions[2].height = 60

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap_align
            status_cell       = row[2]
            status_cell.fill  = data_fill
            status_cell.font  = data_font

    return buf.getvalue()
